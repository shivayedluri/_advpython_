from openpyxl import Workbook
wb=Workbook()
sheet=wb.active
sheet.title="HCL"
j=0
for i in range(10,20,30):
    j+=1
    sheet.cell(row=1,column=j).value=i

for row in sheet.iter_rows(min_row=1,max_row=5,max_col=2,min_col=2):
    for r in row:
        r.value=1

wb.save("../data/demopenxlwrite.xlsx")